import os
from typing import List, Dict, Any

from flask import Flask, jsonify, request, render_template, abort
from openpyxl import Workbook
from supabase import create_client, Client
from dotenv import load_dotenv

APP_DIR = os.path.dirname(__file__)
XLSX_PATH = os.path.join(APP_DIR, "segments.xlsx")


load_dotenv(dotenv_path=os.path.join(APP_DIR, ".env"))

# ---- Supabase setup ----
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")  # Preferably service_role key
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "segments")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ---- Flask app setup ----
app = Flask(__name__, template_folder=os.path.join(APP_DIR, "templates"))

# In-memory pending labels: {segment_id: label}
pending_labels: Dict[int, str] = {}


def load_segments_from_db() -> List[Dict[str, Any]]:
	"""
	Load only segments whose label is NULL or empty string from `segment_images`.
	Apply pending_labels overrides if present.
	"""
	segments: List[Dict[str, Any]] = []
	try:
		# Filter: label IS NULL OR label = ''
		response = (
			supabase
			.table("segment_images")
			.select("*")
			.or_("label.is.null")
			.execute()
		)
		rows = response.data or []
		for row in rows:
			segment_id = row.get("id")
			url = row.get("url", "")
			label = row.get("label") or ""
			if segment_id in pending_labels:
				label = pending_labels[segment_id]
			segments.append({
				"id": segment_id,
				"segment_name": row.get("segment_name", ""),
				"url": url,
				"label": label
			})
	except Exception as e:
		print(f"❌ Failed to load segments from DB: {e}")
	return segments


@app.route("/")
def index() -> Any:
	return render_template("labeler.html")


@app.get("/api/segments")
def api_segments() -> Any:
	rows = load_segments_from_db()
	return jsonify({"segments": rows})


@app.post("/api/pending")
def api_pending() -> Any:
	data = request.get_json(silent=True) or {}
	segment_id = int(data.get("id", 0))
	label = str(data.get("label", ""))
	if segment_id <= 0:
		return jsonify({"ok": False, "error": "invalid segment_id"}), 400
	pending_labels[segment_id] = label
	return jsonify({"ok": True})


@app.post("/api/commit")
def api_commit() -> Any:
	"""
	Commit pending labels back to the Supabase table.
	Also rebuilds the Excel sheet for offline reference.
	"""
	segments = load_segments_from_db()
	for segment in segments:
		segment_id = segment["id"]
		if segment_id in pending_labels:
			label = pending_labels[segment_id]
			try:
				supabase.table("segment_images").update({"label": label}).eq("id", segment_id).execute()
			except Exception as e:
				print(f"❌ Failed to update segment {segment_id}: {e}")

	wb = Workbook()
	ws = wb.active
	ws.title = "results"
	ws.cell(row=1, column=1, value="segment_name")
	ws.cell(row=1, column=2, value="url")
	ws.cell(row=1, column=3, value="label")
	ws.column_dimensions['A'].width = 40
	ws.column_dimensions['B'].width = 80
	ws.column_dimensions['C'].width = 15

	row_idx = 2
	for segment in segments:
		url = segment["url"]
		label = segment.get("label", "")
		segment_name = segment.get("segment_name", "")
		ws.cell(row=row_idx, column=1, value=segment_name)
		ws.cell(row=row_idx, column=2, value=url)
		ws.cell(row=row_idx, column=3, value=label)
		row_idx += 1

	wb.save(XLSX_PATH)
	pending_labels.clear()
	return jsonify({"ok": True})


if __name__ == "__main__":
	app.run(host="127.0.0.1", port=5000, debug=True)
